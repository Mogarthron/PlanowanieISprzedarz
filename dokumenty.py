from openpyxl import load_workbook
from pandas import read_excel, DataFrame
from os import listdir, path


path_to_dok = ".\Dane\Zamowienia\Dok\Dok0021.xlsx"

wb = load_workbook(path_to_dok)

dok_DataFrame = read_excel(path_to_dok, wb.sheetnames[-1])
dok_DataFrame.drop(
    [
        "Miasto",
        "Data realizacji",
        "Unnamed: 7",
        "Numer u kontrahenta",
    ],
    axis=1,
    inplace=True,
)

stan = [x for x in dok_DataFrame["Stan"].unique() if x[0] != "Z"]

filter1 = dok_DataFrame["Stan"] == stan[0]
filter2 = dok_DataFrame["Stan"] == stan[1]
filter3 = dok_DataFrame["Stan"] == stan[2]

nie_zrealizowane = filter1 | filter2 | filter3

dok_DataFrame = dok_DataFrame.loc[nie_zrealizowane]

# print(dok_DataFrame)
